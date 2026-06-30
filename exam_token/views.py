import csv
import io
from decimal import Decimal
from pathlib import Path
from .qr_tracking import QRTrackingError
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import Group, User
from collections import defaultdict
from django.db.models import Count, Q, Prefetch
from django.http import FileResponse, HttpResponse, HttpResponseForbidden
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.http import require_http_methods
from openpyxl import load_workbook
from django.core.paginator import Paginator
from .permissions import filter_visible_phien_thi, home_url_name_for_user
from django.utils.dateparse import parse_date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from .qr_tracking import format_print_date_prefix
from datetime import date, datetime, timedelta
from django.db import transaction
from .models import ChamDiemBaiThi, PhieuCham, PhieuChamChiTiet, DoiChieuChamDiem, GanTokenBaiThi
from .scoring_utils import parse_smart_score
from .constants import (
    ASSIGN_CANCELED,
    EXAM_MINUTE_CHOICES,
    QR_STATUS_CHOICES,
    QR_UNUSED,
    ROLE_ADMIN,
    ROLE_MANAGER,
    ROLE_MARKER,
    ROLE_PROCTOR,
    SESSION_COLLECTED,
    SESSION_FINISHED,
    SESSION_RUNNING,
    LATE_ASSIGN_GRACE_MINUTES,
)
from .forms import PhienThiForm, StaffUserForm
from .models import (
    BaiThiScanDiem,
    CanBo,
    ChamDiemChiTiet,
    ExamSessionLog,
    GanTokenBaiThi,
    HocVien,
    Lop,
    MonHoc,
    PhienThi,
    QRBatch,
    QRTokenBaiThi,
)
from .permissions import can_import_students, can_manage_exam, can_manage_qr, can_manage_users, can_mark_exam, can_view_dashboard
from .services import (
    assign_qr_to_student,
    auto_collect_if_expired,
    cancel_assignment_by_token,
    collect_papers,
    create_qr_batch,
    edit_assignment_sbd,
    finish_session,
    grouped_assignments,
    log_action,
    normalize_status_values,
    session_totals,
)

ACTION_LABELS = {
    'create_session': 'Tạo phiên thi',
    'start_timer': 'Tính giờ',
    'assign_qr': 'Gán QR',
    'edit_sbd': 'Sửa SBD',
    'collect_papers': 'Thu bài',
    'auto_collect': 'Tự thu hết giờ',
    'cancel_qr': 'Hủy bài',
    'finish_session': 'Hoàn thành',
    'print_bien_ban_pdf': 'In biên bản',
    'print_qr_pdf': 'In QR',
    'create_qr_batch': 'Tạo đợt QR',
    'create_user': 'Tạo user',
    'import_students': 'Import thí sinh',
    'score_update': 'Chấm điểm',
}

def permission_denied_page(request):
    return render(request, "exam_token/permission_denied.html", status=403)


def home_by_role(user):
    """
    Trả về URL name phù hợp với quyền của user.
    Hàm này dùng chung cho /, login_view và chuyển trang sau đăng nhập.
    """
    return home_url_name_for_user(user)


def redirect_to_user_home(request, user=None):
    """
    Chuyển user về đúng trang theo quyền.
    Nếu user đã đăng nhập nhưng chưa có role hợp lệ thì trả 403,
    không redirect về login để tránh ERR_TOO_MANY_REDIRECTS.
    """
    user = user or request.user
    destination = home_by_role(user)

    if destination:
        return redirect(destination)

    return permission_denied_page(request)


def root_redirect(request):
    """
    Trang chủ /.
    Không dùng @login_required ở đây để tránh Django tự thêm next=/ gây vòng lặp
    khi cấu hình LOGIN_URL hoặc URL login bị sai.
    """
    if not request.user.is_authenticated:
        return redirect("login")

    return redirect_to_user_home(request)


def user_in_groups(user, group_names):
    """
    Hàm tương thích cho các view chấm điểm đang gọi require_groups.
    Kiểm tra cả superuser, Django Group và can_bo.vai_tro.
    """
    if not user or not user.is_authenticated:
        return False

    if isinstance(group_names, str):
        group_names = [group_names]

    try:
        from .permissions import user_has_any_group
        return user_has_any_group(user, group_names)
    except Exception:
        if user.is_superuser:
            return True
        return user.groups.filter(name__in=group_names).exists()


def require_groups(request, *group_names):
    if len(group_names) == 1 and isinstance(group_names[0], (list, tuple, set)):
        group_names = list(group_names[0])
    else:
        group_names = list(group_names)

    return user_in_groups(request.user, group_names)

def _user_display(user):
    if not user:
        return ""

    full_name = ""
    try:
        full_name = (user.get_full_name() or "").strip()
    except Exception:
        full_name = ""

    username = getattr(user, "username", "") or ""

    if full_name and username:
        return f"{full_name} ({username})"

    return full_name or username

def login_view(request):
    """
    Login an toàn:
    - Nếu đã đăng nhập và vào /dang-nhap/?switch=1 thì logout để đổi user.
    - Nếu đã đăng nhập bình thường thì chuyển đúng trang theo role.
    - Sau POST login thành công thì chuyển đúng trang theo role.
    """
    if request.user.is_authenticated:
        if request.GET.get("switch") == "1":
            logout(request)
            return redirect("login")

        return redirect_to_user_home(request)

    if request.method == "POST":
        username = request.POST.get("username", "").strip()
        password = request.POST.get("password", "")
        user = authenticate(request, username=username, password=password)

        if user is None:
            messages.error(request, "Tài khoản hoặc mật khẩu không đúng.")
        elif not user.is_active:
            messages.error(request, "Tài khoản đã bị khóa.")
        else:
            login(request, user)
            request.session.modified = True
            return redirect_to_user_home(request, user)

    return render(request, "exam_token/login.html")


def logout_view(request):
    logout(request)
    return redirect("login")


@can_view_dashboard
def dashboard(request):
    normalize_status_values()
    sessions_qs  = PhienThi.objects.select_related('lop', 'mon').order_by('-created_at')
    sessions_qs = filter_visible_phien_thi(sessions_qs, request.user)
    sessions = sessions_qs[:100]
    stats = {
        'sessions': PhienThi.objects.count(),
        'running': PhienThi.objects.filter(trang_thai=SESSION_RUNNING).count(),
        'collected': PhienThi.objects.filter(trang_thai=SESSION_COLLECTED).count(),
        'finished': PhienThi.objects.filter(trang_thai=SESSION_FINISHED).count(),
        'qr_unused': QRTokenBaiThi.objects.filter(trang_thai=QR_UNUSED).count(),
    }
    return render(request, 'exam_token/dashboard.html', {'sessions': sessions, 'stats': stats})


def parse_date(value):
    if value is None or value == '':
        return None
    if hasattr(value, 'date'):
        return value.date()
    text = str(value).strip()
    for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y']:
        try:
            return timezone.datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def read_uploaded_rows(uploaded_file):
    name = uploaded_file.name.lower()
    if name.endswith('.csv'):
        text = uploaded_file.read().decode('utf-8-sig')
        return list(csv.DictReader(io.StringIO(text)))
    wb = load_workbook(uploaded_file, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else '' for h in rows[0]]
    result = []
    for row in rows[1:]:
        result.append({headers[i]: row[i] if i < len(row) else None for i in range(len(headers))})
    return result


@can_import_students
def import_students(request):
    if request.method == 'POST':
        uploaded = request.FILES.get('file')
        overwrite = request.POST.get('overwrite') == 'on'
        if not uploaded:
            messages.error(request, 'Vui lòng chọn file Excel/CSV.')
            return redirect('import_students')
        rows = read_uploaded_rows(uploaded)
        created = updated = skipped = 0
        for row in rows:
            ma_hv = str(row.get('ma_hoc_vien') or '').strip()
            ho_ten = str(row.get('ho_ten') or '').strip()
            ma_lop = str(row.get('ma_lop') or '').strip()
            ten_lop = str(row.get('ten_lop') or ma_lop).strip()
            if not ma_hv or not ho_ten or not ma_lop:
                skipped += 1
                continue
            lop, _ = Lop.objects.get_or_create(ma_lop=ma_lop, defaults={'ten_lop': ten_lop})
            if lop.ten_lop != ten_lop and ten_lop:
                lop.ten_lop = ten_lop
                lop.save(update_fields=['ten_lop'])
            defaults = {
                'ho_ten': ho_ten,
                'lop': lop,
                'ngay_sinh': parse_date(row.get('ngay_sinh')),
                'gioi_tinh': row.get('gioi_tinh') or None,
                'trang_thai': 'dang_hoc',
            }
            sbd = row.get('so_bao_danh')
            if sbd not in [None, '']:
                defaults['so_bao_danh'] = int(sbd)
            hv, was_created = HocVien.objects.get_or_create(ma_hoc_vien=ma_hv, defaults=defaults)
            if was_created:
                created += 1
            elif overwrite:
                for k, v in defaults.items():
                    setattr(hv, k, v)
                hv.save()
                updated += 1
            else:
                skipped += 1
        log_action(request, 'import_students', None, 'hoc_vien', None, 'Import danh sách thí sinh', new_data={'created': created, 'updated': updated, 'skipped': skipped})
        messages.success(request, f'Import xong: thêm {created}, cập nhật {updated}, bỏ qua {skipped}.')
        return redirect('import_students')
    return render(request, 'exam_token/import_students.html')


@can_manage_exam
def create_session(request):
    initial = {}
    for key in ['ngay_thi', 'phong_thi', 'can_bo_coi_thi_1', 'can_bo_coi_thi_2']:
        if request.GET.get(key):
            initial[key] = request.GET.get(key)
    if request.method == 'POST':
        form = PhienThiForm(request.POST)
        if form.is_valid():
            pt = form.save(commit=False)
            pt.created_by = request.user
            pt.ma_phien_thi = f'{pt.ngay_thi:%Y%m%d}-{pt.phong_thi}-{pt.lop_id}-{pt.mon_id}-{timezone.now():%H%M%S}'
            pt.save()
            log_action(request, 'create_session', pt, 'phien_thi', pt.pk, 'Tạo phiên thi', new_data={'ma_phien_thi': pt.ma_phien_thi})
            messages.success(request, 'Đã tạo phiên thi.')
            if request.POST.get('open_assign') == '1':
                return redirect('assign_scan', pk=pt.pk)
            return redirect('dashboard')
    else:
        form = PhienThiForm(initial=initial)
    return render(request, 'exam_token/create_session.html', {'form': form, 'exam_minutes': EXAM_MINUTE_CHOICES})

def get_assign_lock_at(pt):
    """
    Thời điểm khóa gán thí sinh mới.
    Chỉ tính sau khi đã bấm bắt đầu thi / tính giờ.
    """
    if not pt.started_at:
        return None

    return pt.started_at + timedelta(minutes=LATE_ASSIGN_GRACE_MINUTES)


def is_new_assignment_locked(pt):
    """
    Khóa gán thí sinh mới nếu:
    - đã thu bài
    - đã hoàn thành
    - hoặc đã quá 15 phút sau started_at
    """
    if pt.finished_at:
        return True

    if pt.collected_at:
        return True

    if pt.trang_thai in [SESSION_COLLECTED, SESSION_FINISHED, "da_thu", "hoan_thanh"]:
        return True

    lock_at = get_assign_lock_at(pt)

    if not lock_at:
        return False

    return timezone.now() >= lock_at


def assign_grace_seconds_left(pt):
    lock_at = get_assign_lock_at(pt)

    if not lock_at:
        return None

    return max(0, int((lock_at - timezone.now()).total_seconds()))

@can_manage_exam
def assign_scan(request, pk):
    pt = get_object_or_404(PhienThi.objects.select_related('lop', 'mon'), pk=pk)

    auto_collect_if_expired(request, pt)
    pt.refresh_from_db()

    is_finished = pt.trang_thai == SESSION_FINISHED or pt.trang_thai == "hoan_thanh" or pt.finished_at is not None
    is_collected = pt.trang_thai == SESSION_COLLECTED or pt.trang_thai == "da_thu" or pt.collected_at is not None

    assign_lock_at = get_assign_lock_at(pt)
    assign_locked = is_new_assignment_locked(pt)
    assign_seconds_left = assign_grace_seconds_left(pt)

    if request.method == 'POST':
        action = request.POST.get('action')

        if is_finished:
            messages.error(request, "Phiên thi đã hoàn thành. Chỉ được phép in biên bản.")
            return redirect("assign_scan", pk=pt.id)

        try:
            if action == 'assign':
                assignment = assign_qr_to_student(
                    request,
                    pt,
                    request.POST.get('token'),
                    request.POST.get('so_bao_danh')
                )
                messages.success(
                    request,
                    f'Đã gán QR cho {assignment.hoc_vien.ho_ten} - tờ {assignment.so_to}.'
                )

            elif action == 'lock_time':
                from .services import start_exam_timer
                start_exam_timer(request, pt)

                messages.success(
                    request,
                    f'Đã bắt đầu tính giờ làm bài. '
                    f'Thí sinh đi trễ vẫn được gán bài trong {LATE_ASSIGN_GRACE_MINUTES} phút đầu.'
                )

            elif action == 'collect_papers':
                collect_papers(request, pt)
                messages.success(request, 'Đã thu bài và khóa ô gán dữ liệu.')

            elif action == 'cancel_by_token':
                assignment = cancel_assignment_by_token(
                    request,
                    pt,
                    request.POST.get('cancel_token'),
                    request.POST.get('cancel_reason', '')
                )
                messages.success(request, f'Đã hủy bài của {assignment.hoc_vien.ho_ten}.')

            elif action == 'edit_sbd':
                edit_assignment_sbd(
                    request,
                    request.POST.get('assignment_id'),
                    request.POST.get('edit_token'),
                    request.POST.get('edit_so_bao_danh')
                )
                messages.success(request, 'Đã sửa SBD.')

            elif action == "finish_session":
                if pt.trang_thai != "da_thu":
                    GanTokenBaiThi.objects.filter(
                        phien_thi=pt,
                        trang_thai="da_gan",
                    ).update(
                        trang_thai="da_thu",
                        thoi_gian_thu=timezone.now(),
                    )

                    QRTokenBaiThi.objects.filter(
                        assignments__phien_thi=pt,
                        trang_thai="da_gan",
                    ).update(
                        trang_thai="da_thu",
                        collected_at=timezone.now(),
                    )

                    pt.collected_at = timezone.now()

                pt.trang_thai = "hoan_thanh"
                pt.finished_at = timezone.now()
                pt.save(update_fields=["trang_thai", "collected_at", "finished_at"])

                log_action(
                    request,
                    "finish_session",
                    pt,
                    "phien_thi",
                    pt.pk,
                    "Hoàn thành việc coi thi, khóa toàn bộ chức năng của phiên thi.",
                )

                messages.success(request, "Đã hoàn thành việc coi thi. Phiên thi đã bị khóa.")
                logout(request)
                return redirect("login")

        except Exception as exc:
            messages.error(request, str(exc))

        return redirect('assign_scan', pk=pt.pk)

    grouped_rows = grouped_assignments(pt)
    totals = session_totals(pt)

    cancel_payload = [
        {
            'token': a.token.token,
            'ho_ten': a.hoc_vien.ho_ten,
            'sbd': a.hoc_vien.so_bao_danh,
            'ma_hoc_vien': a.hoc_vien_id,
            'so_to': a.so_to,
            'trang_thai': a.trang_thai,
        }
        for a in GanTokenBaiThi.objects
        .filter(phien_thi=pt)
        .exclude(trang_thai=ASSIGN_CANCELED)
        .select_related('hoc_vien', 'token')
        .order_by('hoc_vien__so_bao_danh')
    ]

    seconds_left = 0
    if pt.ended_at and pt.trang_thai == SESSION_RUNNING:
        seconds_left = max(0, int((pt.ended_at - timezone.now()).total_seconds()))

    ghep_url = reverse('create_session') + (
        f'?ngay_thi={pt.ngay_thi:%Y-%m-%d}'
        f'&phong_thi={pt.phong_thi}'
        f'&can_bo_coi_thi_1={pt.can_bo_coi_thi_1}'
        f'&can_bo_coi_thi_2={pt.can_bo_coi_thi_2}'
    )

    return render(request, 'exam_token/assign_scan.html', {
        'pt': pt,
        'grouped_rows': grouped_rows,
        'collect_rows': grouped_rows,
        'total_assigned_students': totals['total_students'],
        'total_active_sheets': totals['total_sheets'],
        'cancel_payload': cancel_payload,

        'is_finished': is_finished,
        'is_collected': is_collected,

        # Giữ tên cũ để template ít phải sửa.
        # Nhưng ý nghĩa mới: chỉ khóa sau 15 phút, không khóa ngay khi bắt đầu thi.
        'is_time_locked': assign_locked,
        'is_assign_locked': assign_locked,
        'assign_lock_at': assign_lock_at,
        'assign_grace_seconds_left': assign_seconds_left,
        'assign_grace_minutes': LATE_ASSIGN_GRACE_MINUTES,

        'seconds_left': seconds_left,
        'ghep_url': ghep_url,
    })

def _register_pdf_fonts():
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    normal = getattr(settings, 'VIETNAMESE_PDF_FONT_PATH', '')
    bold = getattr(settings, 'VIETNAMESE_PDF_BOLD_FONT_PATH', '')
    font_name = 'TimesNewRomanVN'
    bold_name = 'TimesNewRomanVNBold'
    try:
        if normal and Path(normal).exists():
            pdfmetrics.registerFont(TTFont(font_name, normal))
        else:
            font_name = 'Helvetica'
        if bold and Path(bold).exists():
            pdfmetrics.registerFont(TTFont(bold_name, bold))
        else:
            bold_name = 'Helvetica-Bold'
    except Exception:
        font_name = 'Helvetica'
        bold_name = 'Helvetica-Bold'
    return font_name, bold_name

def bien_ban_full_class_rows(pt):
    """
    Trả về toàn bộ danh sách học viên trong lớp của phiên thi.
    - Có bài: ghi số tờ.
    - Không có bài: ghi chú Vắng / Không dự thi.
    - Có bài bị hủy: ghi chú thêm số tờ đã hủy.
    """
    students = (
        HocVien.objects
        .filter(lop=pt.lop)
        .order_by("so_bao_danh", "ho_ten", "ma_hoc_vien")
    )

    assignments = (
        GanTokenBaiThi.objects
        .filter(phien_thi=pt)
        .select_related("hoc_vien", "token")
        .order_by("hoc_vien__so_bao_danh", "so_to", "id")
    )

    by_student = {}
    for a in assignments:
        by_student.setdefault(a.hoc_vien_id, []).append(a)

    rows = []
    total_present_students = 0
    total_active_sheets = 0
    total_absent_students = 0
    total_canceled_sheets = 0

    for hv in students:
        hv_assignments = by_student.get(hv.pk, [])

        active_assignments = [
            a for a in hv_assignments
            if a.trang_thai not in [ASSIGN_CANCELED, "huy", "da_huy", "cancelled"]
        ]

        canceled_assignments = [
            a for a in hv_assignments
            if a.trang_thai in [ASSIGN_CANCELED, "huy", "da_huy", "cancelled"]
        ]

        tong_so_to = len(active_assignments)
        ghi_chu_parts = []

        if tong_so_to > 0:
            total_present_students += 1
            total_active_sheets += tong_so_to
        else:
            total_absent_students += 1
            ghi_chu_parts.append("Vắng")

        if canceled_assignments:
            total_canceled_sheets += len(canceled_assignments)
            ghi_chu_parts.append(f"Hủy {len(canceled_assignments)} tờ")

        rows.append({
            "ma_hoc_vien": hv.ma_hoc_vien,
            "sbd": hv.so_bao_danh,
            "ho_ten": hv.ho_ten,
            "ngay_sinh": hv.ngay_sinh,
            "tong_so_to": tong_so_to if tong_so_to > 0 else "",
            "ghi_chu": "; ".join(ghi_chu_parts),
        })

    summary = {
        "total_class_students": len(rows),
        "total_present_students": total_present_students,
        "total_absent_students": total_absent_students,
        "total_active_sheets": total_active_sheets,
        "total_canceled_sheets": total_canceled_sheets,
    }

    return rows, summary

@can_manage_exam
def bien_ban_thi_pdf(request, pk):
    import io
    import math
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.pdfgen import canvas
    from django.shortcuts import get_object_or_404
    from django.http import FileResponse

    pt = get_object_or_404(
        PhienThi.objects.select_related("lop", "mon"),
        pk=pk
    )

    rows, totals = bien_ban_full_class_rows(pt)

    font, bold = _register_pdf_fonts()

    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    x0 = 12 * mm
    col_w = [10, 12, 60, 25, 18, 30, 30]
    headers = ["STT", "SBD", "Họ và tên", "Ngày sinh", "Số tờ", "Chữ ký", "Ghi chú"]
    row_h = 7.5 * mm

    # Tọa độ chuẩn
    header_top_y = height - 22 * mm
    info_y = header_top_y - 35 * mm
    table_y = info_y - 6 * mm
    footer_y = 12 * mm

    # Trang chỉ có bảng có thể chứa nhiều dòng hơn
    max_rows_table_page = int((table_y - 18 * mm) // row_h) - 1

    # Khoảng tối thiểu cần để vẽ tổng + chữ ký
    min_space_for_summary = 80 * mm

    def center_text(text, x, y, size=11, b=False):
        c.setFont(bold if b else font, size)
        c.drawCentredString(x, y, str(text or ""))

    def left_text(text, x, y, size=11, b=False):
        c.setFont(bold if b else font, size)
        c.drawString(x, y, str(text or ""))

    def draw_page_header(page_no, total_pages):
        c.setStrokeColor(colors.black)
        c.setFillColor(colors.black)

        y = header_top_y

        center_text("BỘ CÔNG AN", 55 * mm, y, 11)
        center_text("TRƯỜNG ĐẠI HỌC CSND", 55 * mm, y - 7 * mm, 11, True)

        center_text("CỘNG HOÀ XÃ HỘI CHỦ NGHĨA VIỆT NAM", 145 * mm, y, 11, True)
        center_text("Độc lập – Tự do – Hạnh phúc", 145 * mm, y - 7 * mm, 11, True)

        center_text("BIÊN BẢN GIAO NỘP BÀI THI", width / 2, y - 22 * mm, 15, True)

        left_text(f"Môn: {pt.mon.ten_mon}", 14 * mm, info_y, 11, True)
        left_text(f"Lớp: {pt.lop.ten_lop}", 82 * mm, info_y, 11, True)

        ngay_thi = pt.ngay_thi.strftime("%d/%m/%Y") if pt.ngay_thi else ""
        left_text(f"Ngày thi: {ngay_thi}", 150 * mm, info_y, 11, True)

        center_text(f"Trang {page_no}/{total_pages}", width - 25 * mm, footer_y, 9)

    def draw_table_header():
        c.setLineWidth(0.6)

        x = x0
        for w in col_w:
            c.rect(x, table_y - row_h, w * mm, row_h)
            x += w * mm

        x = x0
        c.setFont(bold, 10)
        for h, w in zip(headers, col_w):
            c.drawCentredString(x + w * mm / 2, table_y - 5.1 * mm, h)
            x += w * mm

    def draw_table_rows(page_rows, start_index):
        c.setFont(font, 11)

        for idx, r in enumerate(page_rows):
            yrow = table_y - (idx + 2) * row_h

            x = x0
            for w in col_w:
                c.rect(x, yrow, w * mm, row_h)
                x += w * mm

            ngay_sinh = ""
            if r.get("ngay_sinh"):
                ngay_sinh = r["ngay_sinh"].strftime("%d/%m/%Y")

            cells = [
                str(start_index + idx + 1),
                str(r.get("sbd") or ""),
                str(r.get("ho_ten") or ""),
                ngay_sinh,
                str(r.get("tong_so_to") or ""),
                "",
                str(r.get("ghi_chu") or ""),
            ]

            x = x0
            for text, w in zip(cells, col_w):
                text = str(text or "")
                if w <= 18:
                    c.drawCentredString(x + w * mm / 2, yrow + 2.4 * mm, text[:10])
                else:
                    c.drawString(x + 2 * mm, yrow + 2.4 * mm, text[:36])
                x += w * mm

        # Trả về y cuối bảng
        return table_y - (len(page_rows) + 1) * row_h

    def draw_summary_and_signatures(start_y):
        """
        start_y là vị trí ngay dưới bảng.
        Phần tổng + chữ ký sẽ đi theo start_y, không cố định cứng 1 trang.
        """
        bottom = start_y - 8 * mm

        c.setFont(font, 11)
        c.drawString(12 * mm, bottom, f'- Tổng số thí sinh trong danh sách: {totals["total_class_students"]}')
        c.drawString(12 * mm, bottom - 6 * mm, f'- Số thí sinh dự thi/có bài: {totals["total_present_students"]}')
        c.drawString(12 * mm, bottom - 12 * mm, f'- Số thí sinh vắng/không dự thi: {totals["total_absent_students"]}')
        c.drawString(12 * mm, bottom - 18 * mm, f'- Tổng số tờ bài thi đã thu: {totals["total_active_sheets"]}')

        sig_y = bottom - 34 * mm

        # Để 3 ô chữ ký cùng một hàng, tránh bị tụt quá sâu khi danh sách dài
        center_text("CÁN BỘ COI THI 1", 38 * mm, sig_y, 10, True)
        center_text("(Ký và ghi rõ họ tên)", 38 * mm, sig_y - 6 * mm, 9)
        center_text(pt.can_bo_coi_thi_1 or "", 38 * mm, sig_y - 30 * mm, 10)

        center_text("CÁN BỘ COI THI 2", 105 * mm, sig_y, 10, True)
        center_text("(Ký và ghi rõ họ tên)", 105 * mm, sig_y - 6 * mm, 9)
        center_text(pt.can_bo_coi_thi_2 or "", 105 * mm, sig_y - 30 * mm, 10)

        center_text("CÁN BỘ NHẬN BÀI THI", 170 * mm, sig_y, 10, True)
        center_text("(Ký và ghi rõ họ tên)", 170 * mm, sig_y - 6 * mm, 9)

    # Chia dòng theo trang bảng
    pages_rows = []
    for i in range(0, len(rows), max_rows_table_page):
        pages_rows.append(rows[i:i + max_rows_table_page])

    if not pages_rows:
        pages_rows = [[]]

    # Kiểm tra trang cuối có đủ chỗ cho phần tổng + chữ ký không
    last_rows = pages_rows[-1]
    last_table_bottom = table_y - (len(last_rows) + 1) * row_h
    need_extra_signature_page = last_table_bottom < min_space_for_summary

    total_pages = len(pages_rows) + (1 if need_extra_signature_page else 0)

    # Vẽ các trang bảng
    global_index = 0
    for page_index, page_rows in enumerate(pages_rows, start=1):
        draw_page_header(page_index, total_pages)
        draw_table_header()
        table_bottom = draw_table_rows(page_rows, global_index)

        is_last_table_page = page_index == len(pages_rows)

        if is_last_table_page and not need_extra_signature_page:
            draw_summary_and_signatures(table_bottom)

        global_index += len(page_rows)

        if page_index < total_pages:
            c.showPage()

    # Nếu trang cuối không đủ chỗ thì tạo thêm trang riêng cho tổng + chữ ký
    if need_extra_signature_page:
        page_no = total_pages
        # draw_page_header(page_no, total_pages)

        c.setFont(bold, 12)
        c.drawString(12 * mm, table_y + 50 * mm, "Tổng hợp giao nộp bài thi")

        draw_summary_and_signatures(table_y + 48 * mm)

    c.save()
    buffer.seek(0)

    log_action(
        request,
        "print_report_pdf",
        pt,
        "phien_thi",
        pt.pk,
        "In biên bản giao nộp PDF"
    )

    return FileResponse(
        buffer,
        filename=f"bien_ban_{pt.ma_phien_thi}.pdf",
        content_type="application/pdf"
    )


@can_manage_qr
def qr_batches(request):
    today = date.today()

    if request.method == 'POST':
        ma_dot = request.POST.get('ma_dot', '').strip()
        # noi_dung_bat_dau = request.POST.get('noi_dung_bat_dau', '').strip()
        noi_dung_bat_dau = "DHCSND_000001"
        ngay_in_raw = request.POST.get("ngay_in", "").strip()

        if ngay_in_raw:
            ngay_in = datetime.strptime(ngay_in_raw, "%Y-%m-%d").date()
        else:
            ngay_in = date.today()

        try:
            so_luong = int(request.POST.get('so_luong') or 0)
        except ValueError:
            so_luong = 0

        if not ma_dot or so_luong <= 0 or not ngay_in:
            messages.error(request, 'Vui lòng nhập mã lô, ngày in và số lượng QR hợp lệ.')
        elif QRBatch.objects.filter(ma_dot=ma_dot).exists():
            messages.error(request, 'Mã lô QR đã tồn tại.')
        else:
            try:
                batch = create_qr_batch(
                    ma_dot=ma_dot,
                    so_luong=so_luong,
                    user=request.user,
                    noi_dung_bat_dau=noi_dung_bat_dau,
                    ngay_in=ngay_in,
                )

                log_action(
                    request,
                    'create_qr_batch',
                    None,
                    'qr_batch',
                    batch.pk,
                    'Tạo lô in QR tracking v2',
                    new_data={
                        'ma_dot': ma_dot,
                        'so_luong': so_luong,
                        'ngay_in': str(batch.ngay_in),
                        'ngay_in_prefix': batch.ngay_in_prefix,
                        'noi_dung_bat_dau': batch.noi_dung_bat_dau,
                        'noi_dung_ket_thuc': batch.noi_dung_ket_thuc,
                    },
                )

                messages.success(
                    request,
                    f'Đã tạo {so_luong} QR cho lô {ma_dot} '
                    f'({batch.noi_dung_bat_dau} → {batch.noi_dung_ket_thuc}). '
                    f'Prefix ngày in: {batch.ngay_in_prefix}.'
                )
            except QRTrackingError as exc:
                messages.error(request, str(exc))
            except Exception as exc:
                messages.error(request, f'Lỗi tạo QR: {exc}')

        return redirect('qr_batches')

    batches = QRBatch.objects.annotate(total=Count('tokens')).order_by('-created_at')
    return render(request, 'exam_token/qr_batches.html', {
        'batches': batches,
        'today': today,
    })


@can_manage_qr
def qr_batch_pdf(request, pk):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.pdfgen import canvas
    import qrcode

    batch = get_object_or_404(QRBatch, pk=pk)
    tokens = list(batch.tokens.order_by('id'))
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    font, bold = _register_pdf_fonts()
    margin = 12 * mm
    card_w = 45 * mm
    card_h = 36 * mm
    cols = 4
    x_gap = 4 * mm
    y_gap = 5 * mm
    c.setFont(bold, 12)
    c.drawString(margin, height - margin, f'Đợt QR: {batch.ma_dot} - {len(tokens)} mã')
    start_y = height - margin - 8 * mm
    for idx, token in enumerate(tokens):
        pos = idx % (cols * 6)
        if idx and pos == 0:
            c.showPage()
            c.setFont(bold, 12)
            c.drawString(margin, height - margin, f'Đợt QR: {batch.ma_dot}')
            start_y = height - margin - 8 * mm
        col = pos % cols
        row = pos // cols
        x = margin + col * (card_w + x_gap)
        y = start_y - row * (card_h + y_gap) - card_h
        c.rect(x, y, card_w, card_h)
        qr_payload = token.qr_link or token.token
        img = qrcode.make(qr_payload)
        img_buf = io.BytesIO()
        img.save(img_buf, format='PNG')
        img_buf.seek(0)
        from reportlab.lib.utils import ImageReader
        c.drawImage(ImageReader(img_buf), x + 3 * mm, y + 9 * mm, 22 * mm, 22 * mm)
        c.setFont(font, 6.5)
        c.drawString(x + 27 * mm, y + 24 * mm, str(token.noi_dung or '')[:24])
        c.drawString(x + 27 * mm, y + 19 * mm, str(token.public_token or token.token)[:24])
        c.setFont(bold, 8)
        c.drawString(x + 27 * mm, y + 12 * mm, 'GIẤY THI QR')
    c.save()
    buffer.seek(0)
    QRTokenBaiThi.objects.filter(batch=batch, printed_at__isnull=True).update(printed_at=timezone.now())
    log_action(request, 'print_qr_pdf', None, 'qr_batch', batch.pk, 'In PDF QR')
    return FileResponse(buffer, filename=f'qr_batch_{batch.ma_dot}.pdf', content_type='application/pdf')


@can_manage_users
def manage_users(request):
    if request.method == 'POST':
        form = StaffUserForm(request.POST)
        if form.is_valid():
            user, can_bo = form.save()
            log_action(request, 'create_user', None, 'can_bo', can_bo.pk, 'Tạo user/cán bộ', new_data={'username': user.username, 'ma_can_bo': can_bo.ma_can_bo})
            messages.success(request, f'Đã tạo tài khoản {user.username}.')
            return redirect('manage_users')
    else:
        form = StaffUserForm()
    users = User.objects.select_related('can_bo').prefetch_related('groups').order_by('username')
    return render(request, 'exam_token/manage_users.html', {'form': form, 'users': users})


@can_mark_exam
def score_list(request):
    scans = BaiThiScanDiem.objects.select_related('phien_thi', 'hoc_vien', 'nguoi_cham').order_by('-created_at')[:200]
    return render(request, 'exam_token/score_list.html', {'scans': scans})


@can_mark_exam
def score_detail(request, pk):
    scan = get_object_or_404(BaiThiScanDiem.objects.select_related('phien_thi__mon', 'phien_thi__lop', 'hoc_vien', 'nguoi_cham'), pk=pk)
    if not scan.chi_tiet.exists():
        ChamDiemChiTiet.objects.bulk_create([
            ChamDiemChiTiet(bai_thi=scan, cau_so=1, noi_dung='Câu 1', diem_toi_da=Decimal('2.00'), diem_dat=Decimal('0.00')),
            ChamDiemChiTiet(bai_thi=scan, cau_so=2, noi_dung='Câu 2', diem_toi_da=Decimal('3.00'), diem_dat=Decimal('0.00')),
            ChamDiemChiTiet(bai_thi=scan, cau_so=3, noi_dung='Câu 3', diem_toi_da=Decimal('5.00'), diem_dat=Decimal('0.00')),
        ])
    details = list(scan.chi_tiet.all())
    if request.method == 'POST':
        action = request.POST.get('action')
        total = Decimal('0')
        for d in details:
            diem = Decimal(request.POST.get(f'diem_{d.id}') or '0')
            if diem < 0 or diem > d.diem_toi_da:
                messages.error(request, f'Điểm câu {d.cau_so} không hợp lệ.')
                return redirect('score_detail', pk=scan.pk)
            d.diem_dat = diem
            d.ghi_chu = request.POST.get(f'ghi_chu_{d.id}', '')
            d.save(update_fields=['diem_dat', 'ghi_chu'])
            total += diem
        scan.diem_chot = total
        scan.nhan_xet = request.POST.get('nhan_xet', '')
        scan.nguoi_cham = getattr(request.user, 'can_bo', None)
        if action == 'confirm':
            scan.trang_thai = 'da_cham'
            scan.thoi_gian_chot = timezone.now()
            msg = 'Đã xác nhận chấm điểm.'
        else:
            scan.trang_thai = 'dang_cham'
            msg = 'Đã lưu nháp điểm.'
        scan.save(update_fields=['diem_chot', 'nhan_xet', 'nguoi_cham', 'trang_thai', 'thoi_gian_chot', 'updated_at'])
        log_action(request, 'mark_exam_confirm' if action == 'confirm' else 'mark_exam_draft', scan.phien_thi, 'bai_thi_scan_diem', scan.pk, msg, new_data={'diem_chot': str(total)})
        messages.success(request, msg)
        return redirect('score_detail', pk=scan.pk)
    return render(request, 'exam_token/score_detail.html', {'scan': scan, 'details': details})


@can_view_dashboard
def logs(request):
    q = request.GET.get('q', '').strip()
    qs = ExamSessionLog.objects.select_related('user', 'can_bo', 'phien_thi').order_by('-created_at')
    if q:
        qs = qs.filter(Q(action__icontains=q) | Q(message__icontains=q) | Q(phien_thi__ma_phien_thi__icontains=q))
    return render(request, 'exam_token/logs.html', {'logs': qs[:300], 'q': q})

@can_view_dashboard
def session_logs(request, pk):
    pt = get_object_or_404(PhienThi.objects.select_related('lop', 'mon', 'created_by'), pk=pk)
    q = request.GET.get('q', '').strip()
    action = request.GET.get('action', '').strip()
    user_id = request.GET.get('user', '').strip()

    base_qs = ExamSessionLog.objects.filter(phien_thi=pt)
    qs = base_qs.select_related('user', 'can_bo', 'phien_thi').order_by('-created_at')

    if q:
        qs = qs.filter(
            Q(action__icontains=q)
            | Q(message__icontains=q)
            | Q(object_type__icontains=q)
            | Q(object_id__icontains=q)
        )
    if action:
        qs = qs.filter(action=action)
    if user_id:
        qs = qs.filter(user_id=user_id)

    paginator = Paginator(qs, 30)
    page_obj = paginator.get_page(request.GET.get('page'))

    actions = list(base_qs.order_by('action').values_list('action', flat=True).distinct())
    users = User.objects.filter(examsessionlog__phien_thi=pt).distinct().order_by('username')

    context = {
        'pt': pt,
        'logs': page_obj.object_list,
        'page_obj': page_obj,
        'q': q,
        'action': action,
        'user_id': user_id,
        'actions': actions,
        'users': users,
        'action_labels': ACTION_LABELS,
        'total_logs': base_qs.count(),
        'total_users': base_qs.exclude(user__isnull=True).values('user_id').distinct().count(),
        'first_log': base_qs.order_by('created_at').first(),
        'last_log': base_qs.order_by('-created_at').first(),
    }
    return render(request, 'exam_token/session_logs.html', context)

@can_manage_qr
def qr_batch_excel(request, pk):
    batch = get_object_or_404(QRBatch, pk=pk)
    tokens = batch.tokens.all().order_by('id')

    wb = Workbook()
    ws = wb.active
    ws.title = 'QR_CODE_LO_IN'

    headers = [
        'lo_in',
        'ngay_in',
        'ngay_in_prefix_10_ky_tu_dau',
        'stt',
        'noi_dung',
        'token_ma_hoa',
        'public_token',
        'QR',
        'trang_thai',
    ]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='1E40AF')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    ngay_in_text = batch.ngay_in.strftime('%Y-%m-%d') if getattr(batch, 'ngay_in', None) else (batch.ngay_in_prefix or '')

    for idx, token in enumerate(tokens, start=1):
        qr_link = token.qr_link or token.token
        ws.append([
            batch.ma_dot,
            ngay_in_text,
            token.ngay_in_prefix or batch.ngay_in_prefix or '',
            idx,
            token.noi_dung or '',
            token.token_ma_hoa or '',
            token.public_token or '',
            qr_link,
            token.trang_thai,
        ])

    widths = {
        'A': 18,
        'B': 14,
        'C': 24,
        'D': 8,
        'E': 22,
        'F': 65,
        'G': 80,
        'H': 95,
        'I': 16,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    safe_ma_dot = str(batch.ma_dot or 'qr_batch').replace('/', '_').replace('\\', '_').replace(' ', '_')
    response['Content-Disposition'] = f'attachment; filename="QR_{safe_ma_dot}.xlsx"'
    wb.save(response)

    log_action(
        request,
        'export_qr_batch_excel',
        None,
        'qr_batch',
        batch.pk,
        f'Xuất Excel nội dung QR lô {batch.ma_dot}',
    )
    return response

def _in_group(user, names):
    if not user or not user.is_authenticated:
        return False
    if user.is_superuser:
        return True
    return user.groups.filter(name__in=names).exists()


def _require_group(user, names):
    return _in_group(user, names + ["admin", "quan_ly_thi"])


def _find_gan_token_by_qr_link(qr_link):
    qr_link = str(qr_link or "").strip()
    if not qr_link:
        raise ValueError("Vui lòng quét QR bài thi.")

    gan = (
        GanTokenBaiThi.objects
        .select_related("phien_thi", "phien_thi__lop", "phien_thi__mon", "token")
        .filter(token__qr_link=qr_link)
        .first()
    )
    if not gan:
        # fallback nếu field token đang lưu qr_link
        gan = (
            GanTokenBaiThi.objects
            .select_related("phien_thi", "phien_thi__lop", "phien_thi__mon", "token")
            .filter(token__token=qr_link)
            .first()
        )
    if not gan:
        raise ValueError("Không tìm thấy bài thi theo QR link đã quét.")

    bad_status = {"chua_gan", "huy", "da_huy"}
    if getattr(gan, "trang_thai", "") in bad_status:
        raise ValueError("Không nhận bài thi chưa gán hoặc đã hủy.")
    if getattr(gan.token, "trang_thai", "") in bad_status:
        raise ValueError("Không nhận QR chưa gán hoặc đã hủy.")

    if getattr(gan, "trang_thai", "") != "da_thu":
        raise ValueError("Chỉ được chấm bài đã thu.")

    return gan


def _get_or_create_bai_cham(gan):
    bai, _ = ChamDiemBaiThi.objects.get_or_create(
        gan_token=gan,
        defaults={
            "phien_thi": gan.phien_thi,
            "token": gan.token,
            "trang_thai": "chua_cham",
        },
    )
    return bai


def _collect_scores_from_post(request):
    cau_list = request.POST.getlist("cau_so[]")
    diem_list = request.POST.getlist("diem[]")
    rows = []
    total = Decimal("0.00")

    for idx, raw_diem in enumerate(diem_list):
        raw_diem = str(raw_diem or "").strip()
        if not raw_diem:
            continue
        try:
            cau_so = int(cau_list[idx] or idx + 1)
        except Exception:
            cau_so = idx + 1
        diem = parse_smart_score(raw_diem)
        rows.append((cau_so, diem))
        total += diem

    if not rows:
        raise ValueError("Vui lòng nhập ít nhất một điểm câu.")
    if total > Decimal("10.00"):
        raise ValueError("Tổng điểm không được vượt quá 10.")

    return rows, total.quantize(Decimal("0.01"))


def _save_phieu_cham(request, bai_cham, lan_cham):
    if PhieuCham.objects.filter(bai_cham=bai_cham, lan_cham=lan_cham, trang_thai="da_nop").exists():
        raise ValueError(f"Bài này đã có phiếu chấm lượt {lan_cham}, không được chấm lại.")

    rows, total = _collect_scores_from_post(request)

    with transaction.atomic():
        phieu = PhieuCham.objects.create(
            bai_cham=bai_cham,
            lan_cham=lan_cham,
            nguoi_cham=request.user,
            tong_diem=total,
            trang_thai="da_nop",
            submitted_at=timezone.now(),
            ghi_chu=request.POST.get("ghi_chu", "").strip(),
        )
        PhieuChamChiTiet.objects.bulk_create([
            PhieuChamChiTiet(phieu_cham=phieu, cau_so=cau, diem=diem)
            for cau, diem in rows
        ])

        if lan_cham == 1:
            bai_cham.trang_thai = "da_cham_gk1"
        elif lan_cham == 2:
            bai_cham.trang_thai = "da_cham_gk2"
        elif lan_cham == 3:
            bai_cham.trang_thai = "hoan_tat"
            bai_cham.diem_chinh_thuc = total
        bai_cham.save(update_fields=["trang_thai", "diem_chinh_thuc", "updated_at"])

    return phieu


@login_required
def cham_diem_gk1(request):
    if not require_groups(request, "giam_khao_1", "admin", "quan_ly_thi"):
        return permission_denied_page(request)
    if not _require_group(request.user, ["giam_khao_1"]):
        return HttpResponseForbidden("Bạn không có quyền chấm lượt 1.")

    context = {"role_title": "Giám khảo 1", "lan_cham": 1, "bai_cham": None}

    if request.method == "POST":
        action = request.POST.get("action")
        qr_link = request.POST.get("qr_link", "").strip()
        try:
            gan = _find_gan_token_by_qr_link(qr_link)
            bai = _get_or_create_bai_cham(gan)

            if action == "scan":
                context.update({"bai_cham": bai, "qr_link": qr_link})
            elif action == "submit_score":
                _save_phieu_cham(request, bai, 1)
                messages.success(request, "Đã nộp phiếu chấm Giám khảo 1.")
                return redirect("cham_diem_gk1")
        except Exception as exc:
            messages.error(request, str(exc))

    return render(request, "exam_token/cham_diem_scan.html", context)


@login_required
def cham_diem_gk2(request):
    if not require_groups(request, "giam_khao_2", "admin", "quan_ly_thi"):
         return permission_denied_page(request)
    if not _require_group(request.user, ["giam_khao_2"]):
        return HttpResponseForbidden("Bạn không có quyền chấm lượt 2.")

    context = {"role_title": "Giám khảo 2", "lan_cham": 2, "bai_cham": None}

    if request.method == "POST":
        action = request.POST.get("action")
        qr_link = request.POST.get("qr_link", "").strip()
        try:
            gan = _find_gan_token_by_qr_link(qr_link)
            bai = _get_or_create_bai_cham(gan)
            if not PhieuCham.objects.filter(bai_cham=bai, lan_cham=1, trang_thai="da_nop").exists():
                raise ValueError("Giám khảo 1 chưa chấm xong bài này.")

            if action == "scan":
                context.update({"bai_cham": bai, "qr_link": qr_link})
            elif action == "submit_score":
                _save_phieu_cham(request, bai, 2)
                messages.success(request, "Đã nộp phiếu chấm Giám khảo 2. Có thể đối chiếu.")
                return redirect("cham_diem_doi_chieu", pk=bai.pk)
        except Exception as exc:
            messages.error(request, str(exc))

    return render(request, "exam_token/cham_diem_scan.html", context)


@login_required
def cham_diem_doi_chieu(request, pk):
    if not _require_group(request.user, ["giam_khao_2"]):
        return HttpResponseForbidden("Bạn không có quyền đối chiếu.")

    bai = get_object_or_404(
        ChamDiemBaiThi.objects.select_related(
            "phien_thi",
            "phien_thi__lop",
            "phien_thi__mon",
        ),
        pk=pk,
    )

    phieu_qs = PhieuCham.objects.select_related("nguoi_cham").prefetch_related("chi_tiet")

    phieu1 = get_object_or_404(
        phieu_qs,
        bai_cham=bai,
        lan_cham=1,
        trang_thai="da_nop",
    )

    phieu2 = get_object_or_404(
        phieu_qs,
        bai_cham=bai,
        lan_cham=2,
        trang_thai="da_nop",
    )

    phieu_thong_nhat = (
        phieu_qs
        .filter(
            bai_cham=bai,
            lan_cham=3,
            trang_thai="da_nop",
        )
        .order_by("-submitted_at", "-id")
        .first()
    )

    do_lech = abs(phieu1.tong_diem - phieu2.tong_diem).quantize(Decimal("0.01"))
    ket_qua = None

    if request.method == "POST":
        with transaction.atomic():
            if do_lech <= Decimal("0.50"):
                ket_qua = "hop_le_tu_dong"
                bai.diem_chinh_thuc = phieu2.tong_diem
                bai.do_lech = do_lech
                bai.trang_thai = "hoan_tat"
                msg = "Đối chiếu hợp lệ. Đã lấy điểm Giám khảo 2 làm điểm chính thức."
            elif do_lech <= Decimal("1.50"):
                ket_qua = "chua_thong_nhat"
                bai.do_lech = do_lech
                bai.trang_thai = "can_cham_thong_nhat"
                msg = "Điểm chưa thống nhất. Cần tạo phiếu chấm thống nhất."
            else:
                ket_qua = "cho_truong_hoi_dong"
                bai.do_lech = do_lech
                bai.trang_thai = "cho_truong_hoi_dong"
                msg = "Điểm lệch trên 1.5. Đã chuyển Trưởng hội đồng xác nhận."

            bai.save(update_fields=["diem_chinh_thuc", "do_lech", "trang_thai", "updated_at"])

            DoiChieuChamDiem.objects.create(
                bai_cham=bai,
                phieu_gk1=phieu1,
                phieu_gk2=phieu2,
                diem_gk1=phieu1.tong_diem,
                diem_gk2=phieu2.tong_diem,
                do_lech=do_lech,
                ket_qua=ket_qua,
                nguoi_doi_chieu=request.user,
            )

        messages.success(request, msg)

        if ket_qua == "chua_thong_nhat":
            return redirect("cham_diem_thong_nhat", pk=bai.pk)

        return redirect("cham_diem_doi_chieu", pk=bai.pk)

    rows = []
    d1 = {x.cau_so: x.diem for x in phieu1.chi_tiet.all()}
    d2 = {x.cau_so: x.diem for x in phieu2.chi_tiet.all()}

    for cau in sorted(set(d1) | set(d2)):
        rows.append({
            "cau_so": cau,
            "diem_gk1": d1.get(cau),
            "diem_gk2": d2.get(cau),
        })

    return render(request, "exam_token/cham_diem_compare.html", {
        "bai_cham": bai,
        "phieu1": phieu1,
        "phieu2": phieu2,
        "phieu_thong_nhat": phieu_thong_nhat,
        "nguoi_cham_gk1": _user_display(phieu1.nguoi_cham),
        "nguoi_cham_gk2": _user_display(phieu2.nguoi_cham),
        "nguoi_cham_thong_nhat": _user_display(phieu_thong_nhat.nguoi_cham) if phieu_thong_nhat else "",
        "rows": rows,
        "do_lech": do_lech,
    })


@login_required
def cham_diem_thong_nhat(request, pk):
    if not _require_group(request.user, ["giam_khao_1", "giam_khao_2"]):
        return HttpResponseForbidden("Bạn không có quyền chấm thống nhất.")

    bai = get_object_or_404(
        ChamDiemBaiThi.objects.select_related(
            "phien_thi",
            "phien_thi__lop",
            "phien_thi__mon",
        ),
        pk=pk,
        trang_thai="can_cham_thong_nhat",
    )

    phieu_qs = PhieuCham.objects.select_related("nguoi_cham").prefetch_related("chi_tiet")

    phieu1 = (
        phieu_qs
        .filter(bai_cham=bai, lan_cham=1, trang_thai="da_nop")
        .order_by("-submitted_at", "-id")
        .first()
    )

    phieu2 = (
        phieu_qs
        .filter(bai_cham=bai, lan_cham=2, trang_thai="da_nop")
        .order_by("-submitted_at", "-id")
        .first()
    )

    if request.method == "POST":
        try:
            _save_phieu_cham(request, bai, 3)
            messages.success(request, "Đã lưu phiếu thống nhất và ghi điểm chính thức.")
            return redirect("cham_diem_doi_chieu", pk=bai.pk)
        except Exception as exc:
            messages.error(request, str(exc))

    return render(request, "exam_token/cham_diem_thong_nhat.html", {
        "bai_cham": bai,
        "phieu1": phieu1,
        "phieu2": phieu2,
        "nguoi_cham_gk1": _user_display(phieu1.nguoi_cham) if phieu1 else "",
        "nguoi_cham_gk2": _user_display(phieu2.nguoi_cham) if phieu2 else "",
        "nguoi_cham_hien_tai": _user_display(request.user),
    })


@login_required
def cham_diem_truong_hoi_dong(request):
    if not _require_group(request.user, ["truong_hoi_dong"]):
        return HttpResponseForbidden("Bạn không có quyền Trưởng hội đồng.")

    if request.method == "POST":
        bai = get_object_or_404(ChamDiemBaiThi, pk=request.POST.get("bai_cham_id"), trang_thai="cho_truong_hoi_dong")
        try:
            diem = parse_smart_score(request.POST.get("diem_chinh_thuc"))
            bai.diem_chinh_thuc = diem
            bai.trang_thai = "hoan_tat"
            bai.ghi_chu = request.POST.get("ghi_chu", "").strip()
            bai.save(update_fields=["diem_chinh_thuc", "trang_thai", "ghi_chu", "updated_at"])
            messages.success(request, "Đã xác nhận điểm chính thức.")
        except Exception as exc:
            messages.error(request, str(exc))
        return redirect("cham_diem_truong_hoi_dong")

    items = (
        ChamDiemBaiThi.objects
        .select_related("phien_thi", "phien_thi__lop", "phien_thi__mon")
        .filter(trang_thai="cho_truong_hoi_dong")
        .order_by("-updated_at")
    )
    return render(request, "exam_token/truong_hoi_dong.html", {"items": items})

@login_required
def bang_diem_lop(request):
    if not require_groups(request, "can_bo_cham_diem", "truong_hoi_dong", "admin", "quan_ly_thi"):
        return permission_denied_page(request)

    phien_thi_id = request.GET.get('phien_thi') or ''

    phien_this = (
        PhienThi.objects
        .select_related('lop', 'mon')
        .order_by('-ngay_thi', '-id')
    )

    selected_phien_thi = None
    rows = []
    max_cau = 2

    if phien_thi_id:
        try:
            selected_phien_thi = phien_this.get(pk=phien_thi_id)
        except PhienThi.DoesNotExist:
            messages.error(request, 'Không tìm thấy phiên thi.')
            return redirect('bang_diem_lop')

        bai_chams = (
            ChamDiemBaiThi.objects
            .filter(phien_thi=selected_phien_thi)
            .select_related(
                'gan_token',
                'gan_token__hoc_vien',
                'phien_thi',
                'phien_thi__lop',
                'phien_thi__mon',
            )
            .prefetch_related(
                Prefetch(
                    'phieu_cham',
                    queryset=PhieuCham.objects.select_related('nguoi_cham').prefetch_related('chi_tiet').order_by('lan_cham', 'id')
                )
            )
            .order_by('gan_token__hoc_vien__so_bao_danh', 'gan_token__hoc_vien__ho_ten')
        )

        for bai in bai_chams:
            hoc_vien = getattr(bai.gan_token, 'hoc_vien', None)
            official_phieu = None

            # Ưu tiên phiếu thống nhất lần 3 nếu có, nếu không lấy phiếu GK2, cuối cùng GK1.
            phieus = list(bai.phieu_cham.all())
            for lan in [3, 2, 1]:
                found = [p for p in phieus if p.lan_cham == lan and p.trang_thai != 'huy']
                if found:
                    official_phieu = found[-1]
                    break

            cau_scores = defaultdict(lambda: '')
            if official_phieu:
                for ct in official_phieu.chi_tiet.all():
                    cau_scores[int(ct.cau_so)] = ct.diem
                    max_cau = max(max_cau, int(ct.cau_so))

            rows.append({
                'hoc_vien': hoc_vien,
                'sbd': getattr(hoc_vien, 'sbd', '') or getattr(hoc_vien, 'so_bao_danh', '') or '',
                'ho_ten': getattr(hoc_vien, 'ho_ten', '') or getattr(hoc_vien, 'full_name', '') or '',
                'ngay_sinh': getattr(hoc_vien, 'ngay_sinh', None),
                'cau_scores': cau_scores,
                'diem_tong': bai.diem_chinh_thuc if bai.diem_chinh_thuc is not None else (official_phieu.tong_diem if official_phieu else ''),
                'nguoi_cham': _user_display(official_phieu.nguoi_cham) if official_phieu else '',
                'lan_cham': official_phieu.lan_cham if official_phieu else '',
                'ghi_chu': getattr(official_phieu, 'ghi_chu', '') if official_phieu else getattr(bai, 'ghi_chu', ''),
                'trang_thai': bai.trang_thai,
            })

    cau_range = list(range(1, max_cau + 1))

    return render(request, 'exam_token/bang_diem_lop.html', {
        'phien_this': phien_this,
        'selected_phien_thi': selected_phien_thi,
        'rows': rows,
        'cau_range': cau_range,
    })

@login_required
def bang_diem_lop_pdf(request):
    if not require_groups(request, "can_bo_cham_diem", "truong_hoi_dong", "giam_khao_1", "giam_khao_2", "admin", "quan_ly_thi"):
        return permission_denied_page(request)

    import io
    from decimal import Decimal
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.pdfgen import canvas
    from reportlab.pdfbase.pdfmetrics import stringWidth

    phien_thi_id = request.GET.get("phien_thi") or ""
    tui_bai = request.GET.get("tui") or ""

    if not phien_thi_id:
        messages.error(request, "Vui lòng chọn phiên thi trước khi xuất PDF.")
        return redirect("bang_diem_lop")

    pt = get_object_or_404(
        PhienThi.objects.select_related("lop", "mon"),
        pk=phien_thi_id
    )

    phieu_qs = (
        PhieuCham.objects
        .select_related("nguoi_cham")
        .prefetch_related("chi_tiet")
        .order_by("lan_cham", "id")
    )

    bai_chams = (
        ChamDiemBaiThi.objects
        .filter(phien_thi=pt)
        .select_related(
            "gan_token",
            "gan_token__hoc_vien",
            "gan_token__token",
            "phien_thi",
            "phien_thi__lop",
            "phien_thi__mon",
        )
        .prefetch_related(
            Prefetch("phieu_cham", queryset=phieu_qs)
        )
        .order_by(
            "gan_token__hoc_vien__so_bao_danh",
            "gan_token__hoc_vien__ho_ten",
            "id",
        )
    )

    def user_display(user):
        if not user:
            return ""
        full_name = ""
        try:
            full_name = (user.get_full_name() or "").strip()
        except Exception:
            full_name = ""
        username = getattr(user, "username", "") or ""
        return full_name or username

    def clean_score(score):
        if score in [None, ""]:
            return ""
        try:
            d = Decimal(str(score)).quantize(Decimal("0.01"))
            text = f"{d:.2f}"
            text = text.rstrip("0").rstrip(".")
            return text
        except Exception:
            return str(score)

    def digit_word(ch):
        mapping = {
            "0": "không",
            "1": "một",
            "2": "hai",
            "3": "ba",
            "4": "bốn",
            "5": "năm",
            "6": "sáu",
            "7": "bảy",
            "8": "tám",
            "9": "chín",
        }
        return mapping.get(ch, ch)

    def score_to_words(score):
        text = clean_score(score)
        if not text:
            return ""

        text = text.replace(",", ".")
        if "." not in text:
            if text == "10":
                return "Mười"
            return digit_word(text).capitalize()

        left, right = text.split(".", 1)
        right = right.rstrip("0")

        if left == "10":
            left_word = "mười"
        else:
            left_word = digit_word(left)

        if not right:
            return left_word.capitalize()

        right_words = " ".join(digit_word(ch) for ch in right)
        return f"{left_word} phẩy {right_words}".capitalize()

    def get_so_phach(bai):
        """
        Ưu tiên lấy số phách từ token.noi_dung.
        Nếu hệ thống sau này có field so_phach thì bổ sung tại đây.
        """
        gan = getattr(bai, "gan_token", None)
        token = None

        if gan:
            token = getattr(gan, "token", None)

        if not token:
            token = getattr(bai, "token", None)

        for attr in ["so_phach", "noi_dung", "token"]:
            value = getattr(token, attr, None) if token else None
            if value:
                return str(value)

        hoc_vien = getattr(gan, "hoc_vien", None) if gan else None
        sbd = getattr(hoc_vien, "so_bao_danh", "") if hoc_vien else ""
        return str(sbd or "")

    rows = []
    gk1_name = ""
    gk2_name = ""

    for bai in bai_chams:
        phieus = list(bai.phieu_cham.all())

        phieu1 = next((p for p in phieus if p.lan_cham == 1 and p.trang_thai != "huy"), None)
        phieu2 = next((p for p in phieus if p.lan_cham == 2 and p.trang_thai != "huy"), None)
        phieu3 = next((p for p in phieus if p.lan_cham == 3 and p.trang_thai != "huy"), None)

        if phieu1 and not gk1_name:
            gk1_name = user_display(phieu1.nguoi_cham)

        if phieu2 and not gk2_name:
            gk2_name = user_display(phieu2.nguoi_cham)

        official_phieu = phieu3 or phieu2 or phieu1

        diem = bai.diem_chinh_thuc
        if diem is None and official_phieu:
            diem = official_phieu.tong_diem

        rows.append({
            "so_phach": get_so_phach(bai),
            "diem_so": clean_score(diem),
            "diem_chu": score_to_words(diem),
            "ghi_chu": getattr(bai, "ghi_chu", "") or "",
        })

    font, bold = _register_pdf_fonts()

    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    margin_x = 11 * mm
    top_y = height - 14 * mm

    def set_font(size=11, b=False):
        c.setFont(bold if b else font, size)

    def center(text, x, y, size=11, b=False):
        set_font(size, b)
        c.drawCentredString(x, y, str(text or ""))

    def left(text, x, y, size=11, b=False):
        set_font(size, b)
        c.drawString(x, y, str(text or ""))

    def fit_text(text, max_width, size=10, b=False):
        text = str(text or "")
        set_font(size, b)
        if stringWidth(text, bold if b else font, size) <= max_width:
            return text
        while text and stringWidth(text + "...", bold if b else font, size) > max_width:
            text = text[:-1]
        return text + "..." if text else ""

    def draw_dotted_line(x1, y1, x2, y2=None):
        if y2 is None:
            y2 = y1

        c.setDash(1, 2)
        c.line(x1, y1, x2, y2)
        c.setDash()

    def draw_header(page_no, total_pages):
        c.setFillColor(colors.black)
        c.setStrokeColor(colors.black)

        center("BỘ CÔNG AN", 50 * mm, top_y, 12)
        center("TRƯỜNG ĐẠI HỌC CSND", 50 * mm, top_y - 7 * mm, 12, True)

        center("CỘNG HOÀ XÃ HỘI CHỦ NGHĨA VIỆT NAM", 145 * mm, top_y, 12, True)
        center("Độc lập - Tự do - Hạnh phúc", 145 * mm, top_y - 7 * mm, 12, True)
        c.line(112 * mm, top_y - 9.5 * mm, 178 * mm, top_y - 9.5 * mm)

        center("PHIẾU TỔNG HỢP ĐIỂM BÀI THI KẾT THÚC HỌC PHẦN", width / 2, top_y - 23 * mm, 14, True)

        y_info = top_y - 36 * mm

        left("Môn:", margin_x + 2 * mm, y_info, 11, True)
        draw_dotted_line(margin_x + 15 * mm, y_info - 1 * mm, margin_x + 95 * mm, y_info - 1 * mm)
        left(fit_text(pt.mon.ten_mon if pt.mon else "", 78 * mm, 10), margin_x + 17 * mm, y_info, 10)

        left("Lớp:", margin_x + 99 * mm, y_info, 11, True)
        draw_dotted_line(margin_x + 110 * mm, y_info - 1 * mm, margin_x + 160 * mm, y_info - 1 * mm)
        left(fit_text(pt.lop.ten_lop if pt.lop else "", 48 * mm, 10), margin_x + 112 * mm, y_info, 10)

        left("Túi:", margin_x + 163 * mm, y_info, 11, True)
        draw_dotted_line(margin_x + 173 * mm, y_info - 1 * mm, width - 10 * mm, y_info - 1 * mm)
        left(str(tui_bai or ""), margin_x + 174 * mm, y_info, 10)

        if total_pages > 1:
            center(f"Trang {page_no}/{total_pages}", width - 22 * mm, 8 * mm, 8)

    def draw_block_table(x, y, data_rows, start_stt, max_rows):
        """
        Vẽ 1 nửa bảng.
        Cấu trúc: STT | Số phách | Điểm bài thi: Bằng số, Bằng chữ | Ghi chú
        """
        col_w = [12 * mm, 18 * mm, 20 * mm, 23 * mm, 19 * mm]
        h1 = 10 * mm
        h2 = 8 * mm
        row_h = 7.1 * mm

        total_w = sum(col_w)
        header_h = h1 + h2

        c.setLineWidth(0.5)

        # Header merged cells
        c.rect(x, y - header_h, col_w[0], header_h)
        c.rect(x + col_w[0], y - header_h, col_w[1], header_h)
        c.rect(x + col_w[0] + col_w[1], y - h1, col_w[2] + col_w[3], h1)
        c.rect(x + col_w[0] + col_w[1], y - header_h, col_w[2], h2)
        c.rect(x + col_w[0] + col_w[1] + col_w[2], y - header_h, col_w[3], h2)
        c.rect(x + col_w[0] + col_w[1] + col_w[2] + col_w[3], y - header_h, col_w[4], header_h)

        center("STT", x + col_w[0] / 2, y - 11 * mm, 10, True)
        center("Số", x + col_w[0] + col_w[1] / 2, y - 7 * mm, 10, True)
        center("phách", x + col_w[0] + col_w[1] / 2, y - 13 * mm, 10, True)
        center("Điểm bài thi", x + col_w[0] + col_w[1] + (col_w[2] + col_w[3]) / 2, y - 7 * mm, 10, True)
        center("Bằng số", x + col_w[0] + col_w[1] + col_w[2] / 2, y - 15 * mm, 9, True)
        center("Bằng chữ", x + col_w[0] + col_w[1] + col_w[2] + col_w[3] / 2, y - 15 * mm, 9, True)
        center("Ghi chú", x + col_w[0] + col_w[1] + col_w[2] + col_w[3] + col_w[4] / 2, y - 11 * mm, 10, True)

        current_y = y - header_h

        for i in range(max_rows):
            row_y = current_y - (i + 1) * row_h

            cx = x
            for w in col_w:
                c.rect(cx, row_y, w, row_h)
                cx += w

            stt = start_stt + i
            row = data_rows[i] if i < len(data_rows) else None

            center(str(stt), x + col_w[0] / 2, row_y + 2.2 * mm, 9)

            if row:
                center(fit_text(row.get("so_phach", ""), col_w[1] - 2 * mm, 8), x + col_w[0] + col_w[1] / 2, row_y + 2.2 * mm, 8)
                center(fit_text(row.get("diem_so", ""), col_w[2] - 2 * mm, 9), x + col_w[0] + col_w[1] + col_w[2] / 2, row_y + 2.2 * mm, 9)
                left(fit_text(row.get("diem_chu", ""), col_w[3] - 2 * mm, 8), x + col_w[0] + col_w[1] + col_w[2] + 1 * mm, row_y + 2.2 * mm, 8)
                left(fit_text(row.get("ghi_chu", ""), col_w[4] - 2 * mm, 8), x + col_w[0] + col_w[1] + col_w[2] + col_w[3] + 1 * mm, row_y + 2.2 * mm, 8)

        return x + total_w, y - header_h - max_rows * row_h

    def draw_footer_and_signatures(y_start):
        left("Thống kê của CBQL:", margin_x + 2 * mm, y_start - 45 * mm, 10)

        box_y = y_start - 41.5 * mm
        labels = ["THCT", "Nhập", "Kiểm dò"]
        x = margin_x + 44 * mm
        for label in labels:
            c.rect(x, box_y - 3.5 * mm, 9 * mm, 5 * mm)
            center(label, x + 4.5 * mm, box_y - 1.7 * mm, 5.5)
            x += 14 * mm

        right_x1 = 103 * mm
        right_x2 = width - 16 * mm
        y = y_start

        center("Ngày ........ tháng ........ năm .............", 150 * mm, y, 11)

        y -= 9 * mm
        center("Họ tên, chữ ký của cán bộ chấm thi 1", 150 * mm, y, 10, True)
        y -= 9 * mm
        draw_dotted_line(right_x1, y, right_x2)
        if gk1_name:
            center(gk1_name, 150 * mm, y + 2 * mm, 9)

        y -= 10 * mm
        center("Họ tên, chữ ký của cán bộ chấm thi 2", 150 * mm, y, 10, True)
        y -= 9 * mm
        draw_dotted_line(right_x1, y, right_x2)
        if gk2_name:
            center(gk2_name, 150 * mm, y + 2 * mm, 9)

        y -= 10 * mm
        center("Họ tên, chữ ký của Lãnh đạo Đơn vị chấm thi", 150 * mm, y, 10, True)
        y -= 9 * mm
        draw_dotted_line(right_x1, y, right_x2)

        left("Mẫu CT.03", margin_x + 1 * mm, 8 * mm, 7)

    rows_per_page = 50
    total_pages = max(1, (len(rows) + rows_per_page - 1) // rows_per_page)

    for page_index in range(total_pages):
        if page_index > 0:
            c.showPage()

        draw_header(page_index + 1, total_pages)

        page_rows = rows[page_index * rows_per_page:(page_index + 1) * rows_per_page]
        left_rows = page_rows[:28]
        right_rows = page_rows[28:50]

        table_top = top_y - 40 * mm
        left_x = margin_x
        right_x = margin_x + 92 * mm

        _, left_bottom = draw_block_table(
            x=left_x,
            y=table_top,
            data_rows=left_rows,
            start_stt=page_index * rows_per_page + 1,
            max_rows=28,
        )

        _, right_bottom = draw_block_table(
            x=right_x,
            y=table_top,
            data_rows=right_rows,
            start_stt=page_index * rows_per_page + 29,
            max_rows=22,
        )

        draw_footer_and_signatures(right_bottom - 6 * mm)

    c.save()
    buffer.seek(0)

    safe_name = str(pt.ma_phien_thi or f"phien_thi_{pt.pk}").replace("/", "_").replace("\\", "_").replace(" ", "_")

    return FileResponse(
        buffer,
        filename=f"phieu_tong_hop_diem_{safe_name}.pdf",
        content_type="application/pdf"
    )